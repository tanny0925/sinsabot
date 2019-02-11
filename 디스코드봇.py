import random
import asyncio
import discord
import openpyxl

client = discord.Client()


@client.event
async def on_ready():
    print("login ")
    print(client.user.name)
    print(client.user.id)
    print("---------------------------------")


@client.event
async def on_message(message):
    if message.content.startswith('ㅎㅇ'):
        await client.send_message(message.channel, "ㅇㅇ")

    if message.content.startswith('굳'):
        await client.send_message(message.channel, "기모딩^^")

    if message.content.startswith('박태현'):
        await client.send_message(message.channel, "21세기 최고의 신사")
    if message.content.startswith('박현석'):
        await client.send_message(message.channel, "아이씨;; ㅡㅡ")

    if message.content.startswith('!명령어'):
        await client.send_message(message.channel, "반응: ㅎㅇ, 굳, 박태현, 박현석  기능(! 붙혀야함): 골라 (a b c), 주사위, 뭐하지, 단어 학습 (A B), 사다리(A B C D/1 1 2 2)  음악 플레이어: !help,!음악,!나가,!들어와,!일시정지,!스킵" )



    if message.content.startswith('!주사위'):
        for i in range(1, 6):
            dice = (int)(random.random() * 6)+1
        await client.send_message(message.channel, (dice))

    if message.content.startswith('!골라'):
        choice = message.content.split(" ")
        choicenumber = random.randint(1, len(choice)-1)
        choiceresult = choice[choicenumber]
        await client.send_message(message.channel, choiceresult)

    if message.content.startswith('!뭐하지'):
        game = "글옵 배그 롤 옵치 레포데 그타 레식"
        gamechoice = game.split(" ")
        gamenumber = random.randint(1, len(gamechoice))
        gameresult = gamechoice[gamenumber-1]
        await client.send_message(message.channel, gameresult)

    if message.content.startswith('!사다리'):
        team = message.content[6:]
        peopleteam = team.split("/")
        people = peopleteam[0]
        team = peopleteam[1]
        person = people.split(" ")
        teamname = team.split(" ")
        random.shuffle(teamname)
        for i in range(0, len(person)):
            await client.send_message(message.channel, person[i] + "-->" + teamname[i])

    if message.content.startswith("!학습"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]:
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await client.send_message(message.channel, "ㅇㅋ 외었따")
                break
        file.save("기억.xlsx")

    if message.content.startswith("!기억") and not message.content.startswith("!기억삭제"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == memory[1]:
                await client.send_message(message.channel, sheet["B" + str(i)].value)
                break

    if message.content.startswith('!기억삭제'):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == str(memory[1]):
                sheet["A" + str(i)].value = "-"
                sheet["B" + str(i)].value = "-"
                await client.send_message(message.channel, "ㅇㅋ 까먹어따")
                file.save("기억.xlsx")
                break





client.run('NTQ0MTY2ODU1MzYzOTE5ODgz.D0HYiw.QLecK5TeSbnRTVK9qViZWEq2L8Y')
